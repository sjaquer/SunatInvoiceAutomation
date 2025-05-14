import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Estructura simplificada del template
columns_metadata = {
    'Item': {
        'required': True,
        'description': 'Número de ítem',
        'example': '1'
    },
    'Product': {
        'required': True,
        'description': 'Descripción del producto (incluir medidas)',
        'example': 'RICE LONG GRAIN 1.00 KG (ARROZ GRANO LARGO 1.00 KG)'
    },
    'Unit': {
        'required': True,
        'description': 'Unidad de medida (BAG, KG, PKT, PCS, etc)',
        'example': 'BAG'
    },
    'Quantity': {
        'required': True,
        'description': 'Cantidad',
        'example': '100'
    },
    'Unit_Price': {
        'required': True,
        'description': 'Precio unitario en dólares',
        'example': '1.440'
    },
    'Total': {
        'required': False,
        'description': 'Calculado automáticamente',
        'example': '144.00'
    }
}

# Datos de ejemplo basados en tu Excel
data = [
    {
        'Item': '1',
        'Product': 'RICE LONG GRAIN 1.00 KG (ARROZ GRANO LARGO 1.00 KG)',
        'Unit': 'BAG',
        'Quantity': '100',
        'Unit_Price': '1.440',
        'Total': '144.00'
    }
]

# Crear DataFrame
df = pd.DataFrame(data)

# Crear archivo Excel con formato
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'comprobantes_template.xlsx')

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Hoja de datos
    df.to_excel(writer, sheet_name='Comprobantes', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Comprobantes']
    
    # Formato para columnas requeridas
    required_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Aplicar formatos
    for idx, col in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        cell = worksheet[f'{column_letter}1']
        
        # Marcar columnas requeridas
        if columns_metadata[col]['required']:
            cell.fill = required_fill
            cell.font = Font(bold=True)
        
        # Ajustar ancho de columna
        max_length = max(
            len(str(cell.value)) for cell in worksheet[column_letter]
        )
        worksheet.column_dimensions[column_letter].width = max_length + 2
    
    # Crear hoja de instrucciones
    instructions_sheet = workbook.create_sheet('Instrucciones', 0)
    
    # Agregar instrucciones
    instructions = [
        ['Campo', 'Requerido', 'Descripción', 'Ejemplo'],
        *[[col, 
           'SÍ' if meta['required'] else 'NO',
           meta['description'],
           meta['example']
        ] for col, meta in columns_metadata.items()]
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        for col_idx, value in enumerate(row, 1):
            cell = instructions_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
            if row_idx == 1:
                cell.font = Font(bold=True)
    
    # Ajustar columnas en hoja de instrucciones
    for column in instructions_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        instructions_sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

print(f"Template creado exitosamente en: {output_path}")
print("\nNotas importantes:")
print("1. Las columnas en amarillo son OBLIGATORIAS")
print("2. Máximo 20 items por comprobante")
print("3. Revisa la hoja 'Instrucciones' para más detalles")

